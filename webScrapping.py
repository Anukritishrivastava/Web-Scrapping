import requests
from bs4 import BeautifulSoup
import pprint
import openpyxl

excel= openpyxl.Workbook()
print(excel.sheetnames)
sheet= excel.active
sheet.title = 'Top news on site'
print(excel.sheetnames)
sheet.append(['Title','Link','votes'])

res = requests.get('https://news.ycombinator.com/news')
res2 = requests.get('https://news.ycombinator.com/news?p=2')
soup = BeautifulSoup(res.text, 'html.parser')
soup2 = BeautifulSoup(res2.text, 'html.parser')

links = soup.select('.titlelink') 
subtext = soup.select('.subtext')
links2 = soup2.select('.titlelink') 
subtext2 = soup2.select('.subtext')

mega_links = links + links2
mega_subtext = subtext + subtext2

def sort_stories_by_votes(hnlist):
  return sorted(hnlist, key= lambda k:k['votes'], reverse=True)

def create_custom_hn(links, subtext):
  hn = []
  for idx, item in enumerate(links):
    title = item.getText()
    href = item.get('href', None)
    vote = subtext[idx].select('.score')
    if len(vote):
      points = int(vote[0].getText().replace(' points', ''))
      if points > 99:
        hn.append({'title': title, 'votes': points, 'link': href})
        sheet.append([title,href,points])
  return sort_stories_by_votes(hn)
 
pprint.pprint(create_custom_hn(mega_links, mega_subtext))

excel.save('topNews.xlsx')
